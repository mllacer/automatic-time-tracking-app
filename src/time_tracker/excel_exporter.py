from __future__ import annotations

import calendar
import logging
import os
import tempfile
from collections import defaultdict
from collections.abc import Iterable
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from time_tracker.models import SessionEvent, WorkSession
from time_tracker.session_builder import SessionBuilder


class ExcelExporter:
    DETAIL_HEADERS = ["Date", "Start", "End", "Duration (hh:mm)", "Duration (decimal hours)", "End Reason"]
    WEEKLY_HEADERS = ["Week Start", "Week End", "Total (hh:mm)", "Total (decimal hours)"]

    def __init__(self, base_dir: Path, logger: logging.Logger | None = None) -> None:
        self._base_dir = base_dir
        self._logger = logger or logging.getLogger(__name__)
        self._builder = SessionBuilder()

    def export(self, events: Iterable[SessionEvent]) -> list[Path]:
        sessions = self._builder.build(events)
        sessions_by_year: dict[int, list[WorkSession]] = defaultdict(list)
        for session in sessions:
            sessions_by_year[session.local_date.year].append(session)

        exported_files: list[Path] = []
        for year, year_sessions in sorted(sessions_by_year.items()):
            target = self._base_dir / f"time-tracking-{year}.xlsx"
            self._write_year_workbook(target=target, year=year, sessions=year_sessions)
            exported_files.append(target)
        return exported_files

    def _write_year_workbook(self, target: Path, year: int, sessions: list[WorkSession]) -> None:
        workbook = Workbook()
        workbook.remove(workbook.active)

        sessions_by_month: dict[int, list[WorkSession]] = defaultdict(list)
        for session in sorted(sessions, key=lambda item: item.start_at):
            sessions_by_month[session.local_date.month].append(session)

        for month, month_sessions in sorted(sessions_by_month.items()):
            month_name = f"{month:02d}-{calendar.month_abbr[month]}"
            detail_sheet = workbook.create_sheet(month_name)
            self._write_detail_sheet(detail_sheet, month_sessions)

            weekly_sheet = workbook.create_sheet(f"{month_name}-weekly")
            self._write_weekly_sheet(weekly_sheet, month_sessions)

        self._save_workbook(workbook, target)
        self._logger.info("Exported workbook for %s to %s", year, target)

    def _write_detail_sheet(self, sheet: Worksheet, sessions: list[WorkSession]) -> None:
        sheet.append(self.DETAIL_HEADERS)
        self._style_header(sheet)
        sheet.freeze_panes = "A2"

        for session in sessions:
            sheet.append(
                [
                    session.local_date,
                    session.start_at.time().replace(tzinfo=None),
                    session.end_at.time().replace(tzinfo=None),
                    format_timedelta_hhmm(session.duration),
                    session.decimal_hours,
                    session.end_reason,
                ]
            )

        sheet.column_dimensions["A"].width = 14
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 12
        sheet.column_dimensions["D"].width = 18
        sheet.column_dimensions["E"].width = 24
        sheet.column_dimensions["F"].width = 18

    def _write_weekly_sheet(self, sheet: Worksheet, sessions: list[WorkSession]) -> None:
        sheet.append(self.WEEKLY_HEADERS)
        self._style_header(sheet)
        sheet.freeze_panes = "A2"

        totals = summarize_by_week(sessions)
        for week_start, duration in sorted(totals.items()):
            week_end = week_start + timedelta(days=6)
            sheet.append(
                [
                    week_start,
                    week_end,
                    format_timedelta_hhmm(duration),
                    round(duration.total_seconds() / 3600, 2),
                ]
            )

        sheet.column_dimensions["A"].width = 14
        sheet.column_dimensions["B"].width = 14
        sheet.column_dimensions["C"].width = 16
        sheet.column_dimensions["D"].width = 22

    @staticmethod
    def _style_header(sheet: Worksheet) -> None:
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    def _save_workbook(self, workbook: Workbook, target: Path) -> None:
        target.parent.mkdir(parents=True, exist_ok=True)
        temp_file = None
        try:
            with tempfile.NamedTemporaryFile(
                dir=target.parent,
                delete=False,
                suffix=".xlsx",
            ) as handle:
                temp_file = Path(handle.name)
            workbook.save(temp_file)
            os.replace(temp_file, target)
        except PermissionError:
            self._logger.warning("Could not write workbook %s because it is in use.", target)
        finally:
            workbook.close()
            if temp_file is not None and temp_file.exists():
                temp_file.unlink(missing_ok=True)


def summarize_by_week(sessions: Iterable[WorkSession]) -> dict[date, timedelta]:
    totals: dict[date, timedelta] = defaultdict(timedelta)
    for session in sessions:
        week_start = session.local_date - timedelta(days=session.local_date.weekday())
        totals[week_start] += session.duration
    return totals


def format_timedelta_hhmm(duration: timedelta) -> str:
    total_minutes = int(duration.total_seconds() // 60)
    hours, minutes = divmod(total_minutes, 60)
    return f"{hours:02d}:{minutes:02d}"
