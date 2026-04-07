from __future__ import annotations

from datetime import date
from datetime import datetime, timedelta, timezone

from openpyxl import load_workbook

from time_tracker.excel_exporter import ExcelExporter
from time_tracker.models import EventType, SessionEvent


LOCAL_TZ = timezone(timedelta(hours=1))


def iso_date(value: date | datetime) -> str:
    return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()


def make_event(event_type: EventType, month: int, day: int, hour: int, minute: int = 0) -> SessionEvent:
    local_time = datetime(2026, month, day, hour, minute, tzinfo=LOCAL_TZ)
    return SessionEvent(
        id=None,
        event_type=event_type,
        occurred_at_utc=local_time.astimezone(timezone.utc),
        local_time=local_time,
        source="test",
    )


def test_exporter_creates_month_and_weekly_sheets(tmp_path) -> None:
    exporter = ExcelExporter(tmp_path)
    events = [
        make_event(EventType.SESSION_START, 1, 5, 8),
        make_event(EventType.LOCK, 1, 5, 13),
        make_event(EventType.UNLOCK, 1, 5, 14),
        make_event(EventType.SHUTDOWN, 1, 5, 18),
    ]

    exported = exporter.export(events)

    assert exported == [tmp_path / "time-tracking-2026.xlsx"]
    workbook = load_workbook(exported[0], data_only=True)
    assert "01-Jan" in workbook.sheetnames
    assert "01-Jan-weekly" in workbook.sheetnames

    detail_sheet = workbook["01-Jan"]
    assert detail_sheet.max_row == 3
    assert iso_date(detail_sheet["A2"].value) == "2026-01-05"
    assert detail_sheet["D2"].value == "05:00"
    assert detail_sheet["E3"].value == 4

    weekly_sheet = workbook["01-Jan-weekly"]
    assert weekly_sheet.max_row == 2
    assert iso_date(weekly_sheet["A2"].value) == "2026-01-05"
    assert weekly_sheet["C2"].value == "09:00"
    assert weekly_sheet["D2"].value == 9


def test_exporter_is_idempotent_for_same_input(tmp_path) -> None:
    exporter = ExcelExporter(tmp_path)
    events = [
        make_event(EventType.SESSION_START, 2, 10, 8),
        make_event(EventType.SHUTDOWN, 2, 10, 12),
    ]

    first_export = exporter.export(events)
    second_export = exporter.export(events)

    assert first_export == second_export

    workbook = load_workbook(first_export[0], data_only=True)
    detail_sheet = workbook["02-Feb"]
    assert detail_sheet.max_row == 2
    assert detail_sheet["D2"].value == "04:00"


def test_exporter_summarizes_multiple_weeks(tmp_path) -> None:
    exporter = ExcelExporter(tmp_path)
    events = [
        make_event(EventType.SESSION_START, 3, 2, 9),
        make_event(EventType.LOCK, 3, 2, 12),
        make_event(EventType.UNLOCK, 3, 9, 10),
        make_event(EventType.SHUTDOWN, 3, 9, 15),
    ]

    exported = exporter.export(events)

    workbook = load_workbook(exported[0], data_only=True)
    weekly_sheet = workbook["03-Mar-weekly"]
    assert weekly_sheet.max_row == 3
    assert weekly_sheet["C2"].value == "03:00"
    assert weekly_sheet["C3"].value == "05:00"
